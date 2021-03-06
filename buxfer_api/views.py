# from django.shortcuts import render
import requests
import urllib.parse
import time
import six
from django.db import DatabaseError
from django.shortcuts import render, redirect
from rest_framework import viewsets
from django.conf import settings
from .serializers import AccountSerializer, TagSerializer, TransactionSerializer
from .models import Account, Tag, Transaction, DollarPrice, TIPOS_TAG_ORD, CAT_PRINCIPALES
from .forms import AccountForm, TagForm, TransactionForm, EditDataForm
from .utils import get_ars_amount, transform_USD_ARS
from django.forms.models import modelformset_factory
import pprint
import re
import math
# from openpyxl import Workbook
from openpyxl import load_workbook
from django.db.models import Q
import datetime
from dateutil.relativedelta import relativedelta

# import pdb
# Create your views here.
requests.packages.urllib3.disable_warnings()
import json
import time


def buxfer_login(request):
    ''' performs the login in the buxfer api and returns the token and a status text '''
    response = requests.get("%s/login?userid=%s" \
                            "&password=%s" % (
                            settings.BUXFER_CONN['url'], settings.BUXFER_CONN['user'], settings.BUXFER_CONN['pass']))

    if response.status_code != 200:
        return (None, get_err_msg(response))

    token = response.json()
    return (token['response']['token'], 'ok')


def buxfer_get_token(request):
    '''  gets the the Buxfer token from the request object or doing a login in buxfer api '''
    status = 'ok'
    token = request.session.get('buxfer_token', None)
    if not token:
        (token, status) = buxfer_login(request)
    return (token, status)


def get_err_msg(response):
    ''' Formats the error message of a requests response'''
    data = response.json()
    err_type = data['error']['type']
    err_msg = data['error']['message']
    return "({}) {} {}".format(response.status_code, err_type, err_msg)


def buxfer_api_fetch(request, resource, page=None):
    '''
    Fetches a resource from the buxfer api
    :param request: django request object
    :param resource:
    :return: dictionary containing the data of the request in json and a string inidcating 'ok' or 'error'
    '''
    attempt_num = 0  # keep track of how many times we've retried
    while attempt_num < settings.BUXFER_CONN['max_retiries']:
        (token, status) = buxfer_get_token(request)
        if not token:
            return {'data': None, 'status': 'Buxfer Login Error: {}'.format(status)}
        url = "%s/%s?token=%s" % (settings.BUXFER_CONN['url'],
                                  resource, token)
        if page:
            url = '{}&page={}'.format(url, page)
        response = requests.get(url, timeout=10)
        print(url)
        if response.status_code == 200:
            data = response.json()
            return {'data': data, 'status': 'ok'}
        else:
            attempt_num += 1
            # You can probably use a logger to log the error here
            time.sleep(5)  # Wait for 5 seconds before re-trying

    err_msg = get_err_msg(response)
    return {'data': None, 'status': 'Error: {}'.format(err_msg)}


def has_changed(record, data):
    for key in data:
        # print(getattr(record, key, None))
        if data[key] != getattr(record, key, None):
            return True
    return False


def save_serializer(actionResponseData, keys, queryset, serializerClass):
    result = []
    serialized_keys = []
    for acc in actionResponseData:
        serializer = None
        # pprint.pprint(acc)
        # si uso many=True en el serializer, graba solo algunos
        if acc['id'] in keys:
            db_rec = queryset[acc['id']]
            if has_changed(db_rec, acc):
                serializer = serializerClass(db_rec, data=acc)
            else:
                result.append('{} not changed'.format(acc['id']))
        else:
            serializer = serializerClass(data=acc)

        if serializer:
            if serializer.is_valid():
                serialized_keys.append(acc['id'])
                embed = serializer.save()
                result.append(embed)
            else:
                result.append(serializer.errors)
    # delete the ones that are not in buxfer anymore
    # to_delete = [x for x in keys if x not in serialized_keys]
    # modelClass.objects.filter(pk__in=to_delete).delete()

    return result


def data_import(request, action, modelClass, serializerClass, formClass=None):
    ''' Imports data from buxfer '''
    form_len = 25
    result = []
    to_delete = None
    formSetClass = modelformset_factory(modelClass, form=formClass, extra=0, can_delete=False)

    if request.method != 'POST':
        data = buxfer_api_fetch(request, action)
        if data['status'] == 'ok':
            queryset = dict([(x.id, x) for x in modelClass.objects.all()])
            keys = list(six.iterkeys(queryset))
            #import pdb  ; pdb.set_trace()

            result = save_serializer(data['data']['response'][action], keys, queryset, serializerClass)

            if action == 'transactions':
                numTransactions = data['data']['response']['numTransactions']
                print(numTransactions)
                for page in range(2, math.ceil(int(numTransactions) / 25) + 1):
                    data = buxfer_api_fetch(request, action, page)
                    if data['data']:
                        result += save_serializer(data['data']['response'][action], keys, queryset, serializerClass)
                        time.sleep(1)
            print(result)
            # import pdb;pdb.set_trace()
            ids = [x.id for x in result if hasattr(x, 'id')]  # not isinstance(x, str)]
            pIds = ids[:form_len]
            if formClass and ids:
                queryset = modelClass.objects.filter(id__in=pIds)
                formset = formSetClass(queryset=queryset)
                qData = zip(list(queryset), list(formset))
                editData = EditDataForm(data={'current_page': 0, 'next_page': 1, 'ids': json.dumps(ids)})
            else:
                formset = None
                qData = None
                editData = None
    else:
        errors = 0
        # import pdb; pdb.set_trace()
        data = {'status': 'No Changes.'}
        if formClass:
            # regex_list=['the', 'an?']
            # regex_list_compiled=[re.compile("^"+i+"$") for i in regex_list]
            # extractddicti= {k:v for k,v in dicti.items() if any (re.match(regex,k) for regex in regex_list_compiled)}
            editData = EditDataForm(request.POST)
            if editData.is_valid():
                ids = json.loads(editData.cleaned_data['ids'])
                pageNum = editData.cleaned_data['current_page']
                nextPage = editData.cleaned_data['next_page']
            else:
                pattern = 'form-[0-9]+-id'
                ids = [value for key, value in request.POST.items() if re.search(pattern, key)]
                pageNum = 0
                nextPage = 0
            offset = pageNum * form_len
            pIds = ids[offset:offset + form_len]
            queryset = modelClass.objects.filter(id__in=pIds)
            formset = formSetClass(request.POST, request.FILES)
            qData = zip(list(queryset), list(formset))
            # import pdb; pdb.set_trace()
            if formset.is_valid():
                # do something with the formset.cleaned_data
                for form in formset:
                    if form.is_valid() and form.changed_data:
                        try:
                            # if form.cleaned_data.get('DELETE') and form.instance.pk:
                            #    form.instance.delete()
                            # else:
                            form.save()
                            data['status'] = "Saved."
                        except DatabaseError as e:
                            data['status'] = "Error: {}".format(e)
                            errors += 1
            if not errors:
                #import pdb;
                #pdb.set_trace()
                offset = nextPage * form_len
                pIds = ids[offset:offset + form_len]
                queryset = modelClass.objects.filter(id__in=pIds)
                if queryset:
                    formset = formSetClass(queryset=queryset)
                    qData = zip(list(queryset), list(formset))
                    editData = EditDataForm(data={'current_page': nextPage, 'next_page': nextPage + 1, 'ids': ids})

    qFieldNames = modelClass().get_fieldNames()
    return render(request, "buxfer_api/data_imported.html", {'status': data['status'], 'data': result, 'action': action,
                                                             'deleted': to_delete, 'qdata': qData,
                                                             'qFieldNames': qFieldNames,
                                                             'formset': formset, 'editData': editData})


def accounts_import(request):
    ''' Imports the accounts'''
    action = 'accounts'
    serializerClass = AccountSerializer
    model = Account
    formSetClass = AccountForm
    return data_import(request, action, model, serializerClass, formSetClass)


def tags_import(request):
    ''' Imports the accounts'''
    action = 'tags'
    serializerClass = TagSerializer
    model = Tag
    formSetClass = TagForm
    return data_import(request, action, model, serializerClass, formSetClass)


def transactions_import(request):
    ''' Imports the accounts'''
    accounts_import(request)
    tags_import(request)
    action = 'transactions'
    serializerClass = TransactionSerializer
    model = Transaction
    formSetClass = TransactionForm
    return data_import(request, action, model, serializerClass, formSetClass)

def clasificar_transacciones(anio, mes, addDiscrecionalidad=False):
    '''Clasifica las transacciones para mostrar con totales'''
    data = {}
    total = 0
    tipos_tag = {}

    if anio and mes:
        transactions = Transaction.objects.filter(date__year=anio, date__month=mes).order_by(
            '-date')
    elif anio:
        transactions = Transaction.objects.filter(date__year=anio).order_by('-date')
    else:
        transactions = Transaction.objects.order_by('-date')

    for (key, value) in TIPOS_TAG_ORD:
        data[value] = {}
        data[value]['data'] = {}
        data[value]['total'] = 0
        tipos_tag[key] = value

    categorias = dict([x for x in CAT_PRINCIPALES])

    # import pdb;pdb.set_trace()
    for transaction in transactions:
        tags = transaction.tags.all()
        if not transaction.cantPesos:
            transaction.cantPesos = get_ars_amount(transaction)
            transaction.save()
        if tags:
            tag = tags[0]
            if tag.tipo_tag and tag.categoria and tag.categoria in categorias:
                tiponame = tipos_tag[tag.tipo_tag]
                catname = categorias[tag.categoria]
                if catname not in data[tiponame]['data']:
                    data[tiponame]['data'][catname] = {}
                    data[tiponame]['data'][catname]['data'] = {}
                    data[tiponame]['data'][catname]['total'] = 0
                if tag.name not in data[tiponame]['data'][catname]['data']:
                    data[tiponame]['data'][catname]['data'][tag.name] = {}
                    data[tiponame]['data'][catname]['data'][tag.name]['data'] = []
                    data[tiponame]['data'][catname]['data'][tag.name]['total'] = 0
                data[tiponame]['data'][catname]['data'][tag.name]['data'].append(transaction)
                if addDiscrecionalidad:
                    if 'discrecionalidad' not in data[tiponame]['data'][catname]['data'][tag.name]:
                        data[tiponame]['data'][catname]['data'][tag.name]['discrecionalidad'] = {}
                    dct = {}
                    if transaction.discrecionalidad:
                        dct['discrecionalidad'] = transaction.discrecionalidad
                    else:
                        dct['discrecionalidad'] = tag.discrecionalidad
                    # dct['tipo_tag'] = tag.tipo_tag
                    dct['tipo_pago'] = transaction.accountId.tipo_pago
                    data[tiponame]['data'][catname]['data'][tag.name]['discrecionalidad'][transaction.id] = dct
                data[tiponame]['data'][catname]['data'][tag.name]['total'] += transaction.cantPesos
                data[tiponame]['data'][catname]['total'] += transaction.cantPesos
                data[tiponame]['total'] += transaction.cantPesos
                total += transaction.cantPesos

    # import pdb; pdb.set_trace()
    show_data = [(x[1], data[x[1]]) for x in TIPOS_TAG_ORD]

    return show_data, total


def cuadro_activos(request, anio=None, mes=None):
    ''' Muestra tabla de activos, pasivos, ingresos y gastos '''

    show_data, total = clasificar_transacciones(anio, mes)
    # pprint.pprint(show_data)

    return render(request, "buxfer_api/cuadro_activos.html", {'data': show_data, 'total': total})


def planilla_estado_financiero(request, anio=None, mes=None):
    '''Graba la planilla de estado financiero'''

    data, total = clasificar_transacciones(anio, mes, True)

    columnas_meses = {1: 'D', 2: 'E', 3: 'F', 4: 'G', 5: 'H', 6: 'I', 7: 'J', 8: 'K', 9: 'L', 10: 'M', 11: 'N', 12: 'O'}
    filas_categorias = {'HOGAR Y VIVIENDA': (39, 2),
                        'AUTOMOVIL Y TRANSPORTE': (40, 3),
                        'ALIMENTOS': (41, 4),
                        'ROPA': (42, 5),
                        'CUIDADO PERSONAL': (43, 6),
                        'CUIDADO DE LA SALUD': (44, 7),
                        'ENTRETENIMIENTO': (45, 8),
                        'REGALOS': (46, 9),
                        'EDUCACION': (47, 10),
                        'VACACIONES': (48, 11),
                        'GASTOS DE NEGOCIOS': (49, 12),
                        'CUIDADO Y DEPENDENCIAS': (50, 13),
                        'INVERSIONES Y AHORROS': (51, 14),
                        'SEGUROS': (52, 15),
                        'ESPIRITUAL': (53, 16),
                        'DEUDAS REPAGADAS': (54, 17),
                        'SERVICIOS': (55, 18),
                        'IMPUESTOS': (56, 19),
                        'LECCIONES APRENDIDAS': (57, 20),
                        'HONORARIOS PAGADOS': (58, 21)
                        }
    workbook = load_workbook(filename="estado_financiero_plantilla.xlsx")
    sheetBalance = workbook.get_sheet_by_name('Balance y Flujo')
    sheetGastos = workbook.get_sheet_by_name('Gastos Detalle')
    # print(workbook.sheetnames) # ['Balance y Flujo', 'Gastos Detalle']
    totales = {}
    transactionLine = 25
    # pdb.set_trace()
    for item in data:
        for categoria, tags in item[1]['data'].items():
            for tag, transacciones in tags['data'].items():
                for transaccion in transacciones['data']:
                    if not transaccion.cantPesos:
                        transaccion.cantPesos = get_ars_amount(transaccion)
                        transaccion.save()
                    sheetGastos["A{}".format(transactionLine)] = transaccion.date  # fecha
                    sheetGastos["B{}".format(transactionLine)] = transaccion.description  # descripcion
                    sheetGastos["C{}".format(transactionLine)] = tags['data'][tag]['discrecionalidad'][transaccion.id][
                        'discrecionalidad']  # F/V/D
                    sheetGastos["D{}".format(transactionLine)] = tags['data'][tag]['discrecionalidad'][transaccion.id][
                        'tipo_pago']  # $/T/C
                    sheetGastos["E{}".format(transactionLine)] = categoria  # cuenta
                    sheetGastos["F{}".format(transactionLine)] = tag  # subcuenta
                    sheetGastos["G{}".format(transactionLine)] = transaccion.cantPesos  # monto

                    transactionLine += 1
            # pdb.set_trace()

            sheetBalance["{}{}".format('C', filas_categorias[categoria][0])] = tags['total']
            sheetGastos["{}{}".format(columnas_meses[mes], filas_categorias[categoria][1])] = tags['total']

    workbook.save(filename="estado_financiero_{}_{}.xlsx".format(anio, mes))

    return render(request, "buxfer_api/planilla_estado_financiero.html")


def get_api_call(ids, **kwargs):
    API_BASE_URL = "https://apis.datos.gob.ar/series/api/"
    kwargs["ids"] = ",".join(ids)
    return "{}{}?{}".format(API_BASE_URL, "series", urllib.parse.urlencode(kwargs))


def fetch_precio_dolar(fecha):
    '''Se trae el precio del dolar oficial mostrador banco nacion de las 15 hs'''
    # fecha: "2017-01-01"

    #           fecha       compra  venta
    # 'data': [['2017-04-03', 15.2, 15.6],
    #          ['2017-04-04', 15.2, 15.6],
    #
    # https://datos.gob.ar/series/api/series/?ids=tc_usd_bna_mc15
    # Tipo de cambio ARS/USD. BNA. Mostrador. Compra. 15hs. (tc_usd_bna_mc15)
    #
    # https://datos.gob.ar/series/api/series/?ids=tc_usd_bna_mv15
    # Tipo de cambio ARS/USD. BNA. Mostrador. Venta. 15hs. (tc_usd_bna_mv15)
    #

    api_call = get_api_call(["tc_usd_bna_mc15", "tc_usd_bna_mv15"], start_date=fecha)
    print(api_call)
    result = requests.get(api_call).json()
    print(result)
    if 'data' in result:
        return result['data']
    return []


def add_precios_dolar(request, modo, anio=0, mes=0, dia=0):
    '''Obtener los precios del dolar oficial'''
    updated = []
    formato_fecha = '%Y-%m-%d'
    hoy = datetime.date.today().strftime(formato_fecha)
    db_dates = [x.purchase_date.strftime(formato_fecha) for x in DollarPrice.objects.all()]

    # modo == 'todos'
    fecha = '2000-01-01'
    if modo == 'actualizar' and db_dates:
        fecha = max(db_dates)
    elif modo == 'fecha':
        fecha = datetime.date(anio, mes, dia).strftime(formato_fecha)

    last_date = fecha
    while fecha < hoy:

        precios = fetch_precio_dolar(fecha)

        if not precios:
            break
        last_date = precios[-1][0]
        if last_date == fecha:
            break
        fecha = last_date

        for linea in precios:
            if linea[0] not in db_dates and (linea[1] or linea[2]):
                dp = DollarPrice(purchase_date=linea[0], buy=linea[1], sell=linea[2])
                dp.save()
                updated.append(linea)

    return render(request, "cotizacion/price_updated.html", {'updated': updated})

def convertir_moneda(request, moneda_origen, moneda_destino, anio, mes, dia, cantidad, impuesto_pais=True):
    conversiones = {'USD_ARS': transform_USD_ARS}
    origen = moneda_origen.upper().strip()
    destino = moneda_destino.upper().strip()
    par = '{}_{}'.format(origen, destino)
    fecha = datetime.date(anio, mes, dia)
    monto = None
    if par in conversiones:
        monto = conversiones[par](cantidad, fecha, impuesto_pais)
    return render(request, "cotizacion/conversion.html", {'monto': monto, 'origen': origen, 'destino': destino,
                                                          'fecha': fecha, 'impuesto_pais': impuesto_pais,
                                                          'cantidad': cantidad})

convertir_moneda